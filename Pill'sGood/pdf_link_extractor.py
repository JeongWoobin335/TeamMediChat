"""
Excel 파일에서 PDF 링크를 추출하고 다운로드하여 내용을 파싱하는 모듈
"""
import os
import re
import requests
from typing import Dict, Optional, List
from langchain_community.document_loaders import PyPDFLoader
from cache_manager import cache_manager
import tempfile
import pandas as pd
from langchain_openai import ChatOpenAI
from concurrent.futures import ThreadPoolExecutor, as_completed

# .xlsx 파일용
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# LLM 초기화 (요약용) - 순환 import 방지를 위해 여기서 직접 초기화
_summary_llm = None

def _get_summary_llm():
    """요약용 LLM을 지연 초기화 (순환 import 방지)"""
    global _summary_llm
    if _summary_llm is None:
        _summary_llm = ChatOpenAI(model="gpt-4o", temperature=0)
    return _summary_llm

def extract_hyperlinks_from_excel(excel_file_path: str, row_index: int, 
                                  column_mapping: Dict[str, str] = None) -> Dict[str, Optional[str]]:
    """
    Excel 파일에서 특정 행의 하이퍼링크를 추출합니다.
    .xls와 .xlsx 파일 모두 지원합니다.
    
    Args:
        excel_file_path: Excel 파일 경로
        row_index: 행 인덱스 (0부터 시작, pandas iterrows 기준)
        column_mapping: {간단한_이름: 실제_컬럼명} 매핑 (예: {'효능': '이 약의 효능은 무엇입니까?'})
                        None이면 자동으로 컬럼명 찾기 시도
    
    Returns:
        {간단한_이름: URL} 딕셔너리
    """
    file_ext = os.path.splitext(excel_file_path)[1].lower()
    url_pattern = r'https?://[^\s]+'
    
    # .xls 파일인 경우 pandas로 처리
    if file_ext == '.xls':
        try:
            df = pd.read_excel(excel_file_path)
            hyperlinks = {}
            
            if column_mapping:
                # 매핑이 제공된 경우
                for simple_name, actual_col_name in column_mapping.items():
                    if not actual_col_name:  # 빈 문자열은 건너뛰기
                        continue
                    if actual_col_name in df.columns:
                        cell_value = df.iloc[row_index][actual_col_name]
                        if pd.notna(cell_value):
                            cell_str = str(cell_value)
                            # URL 패턴 확인
                            match = re.search(url_pattern, cell_str)
                            if match:
                                hyperlinks[simple_name] = match.group(0)
                            else:
                                hyperlinks[simple_name] = None
                        else:
                            hyperlinks[simple_name] = None
                    else:
                        hyperlinks[simple_name] = None
            else:
                # 매핑이 없는 경우 자동으로 찾기
                keywords = {'효능': ['효능', '효과'], '복용법': ['복용', '사용', '용법'], '주의사항': ['주의', '부작용', '이상반응']}
                for simple_name, search_keywords in keywords.items():
                    found = False
                    for col_name in df.columns:
                        if any(keyword in str(col_name) for keyword in search_keywords):
                            cell_value = df.iloc[row_index][col_name]
                            if pd.notna(cell_value):
                                cell_str = str(cell_value)
                                match = re.search(url_pattern, cell_str)
                                if match:
                                    hyperlinks[simple_name] = match.group(0)
                                    found = True
                                    break
                    if not found:
                        hyperlinks[simple_name] = None
            
            return hyperlinks
        
        except Exception as e:
            print(f"⚠️ Excel (.xls) 하이퍼링크 추출 실패: {e}")
            return {}
    
    # .xlsx 파일인 경우 openpyxl로 처리
    elif file_ext == '.xlsx' and HAS_OPENPYXL:
        try:
            wb = load_workbook(excel_file_path, data_only=False)
            ws = wb.active
            
            # 컬럼명을 인덱스로 변환
            header_row = 1  # 첫 번째 행이 헤더라고 가정
            column_indices = {}
            
            if column_mapping:
                # 매핑이 제공된 경우
                for simple_name, actual_col_name in column_mapping.items():
                    if not actual_col_name:  # 빈 문자열은 건너뛰기
                        continue
                    for col_idx, cell in enumerate(ws[header_row], start=1):
                        cell_value = str(cell.value) if cell.value else ""
                        if actual_col_name in cell_value or cell_value == actual_col_name:
                            column_indices[simple_name] = col_idx
                            break
            else:
                # 매핑이 없는 경우 자동으로 찾기
                keywords = {'효능': ['효능', '효과'], '복용법': ['복용', '사용', '용법'], '주의사항': ['주의', '부작용', '이상반응']}
                for simple_name, search_keywords in keywords.items():
                    for col_idx, cell in enumerate(ws[header_row], start=1):
                        cell_value = str(cell.value) if cell.value else ""
                        if any(keyword in cell_value for keyword in search_keywords):
                            column_indices[simple_name] = col_idx
                            break
            
            hyperlinks = {}
            actual_row = row_index + 2  # 헤더 행 다음부터 시작 (1-based index)
            
            for simple_name, col_idx in column_indices.items():
                cell = ws.cell(row=actual_row, column=col_idx)
                
                # 하이퍼링크 확인
                if cell.hyperlink:
                    hyperlinks[simple_name] = cell.hyperlink.target
                elif cell.value and isinstance(cell.value, str):
                    # 셀 값이 URL인지 확인
                    match = re.search(url_pattern, cell.value)
                    if match:
                        hyperlinks[simple_name] = match.group(0)
                    else:
                        hyperlinks[simple_name] = None
                else:
                    hyperlinks[simple_name] = None
            
            wb.close()
            return hyperlinks
        
        except Exception as e:
            print(f"⚠️ Excel (.xlsx) 하이퍼링크 추출 실패: {e}")
            return {}
    
    else:
        print(f"⚠️ 지원하지 않는 파일 형식: {file_ext}")
        return {}


def download_pdf_from_url(url: str, cache_key: str = None) -> Optional[str]:
    """
    URL에서 PDF를 다운로드하여 임시 파일로 저장합니다.
    
    Args:
        url: PDF 다운로드 URL
        cache_key: 캐시 키 (선택사항)
    
    Returns:
        다운로드된 PDF 파일 경로 (실패 시 None)
    """
    if not url or not url.startswith(('http://', 'https://')):
        return None
    
    # 캐시 확인
    if cache_key:
        cached_path = cache_manager.get_pdf_cache(cache_key)
        if cached_path and os.path.exists(cached_path):
            print(f"📂 PDF 캐시 히트: {cache_key}")
            return cached_path
    
    try:
        print(f"📥 PDF 다운로드 중: {url}")
        response = requests.get(url, timeout=30, allow_redirects=True)
        response.raise_for_status()
        
        # Content-Type 확인
        content_type = response.headers.get('Content-Type', '').lower()
        if 'pdf' not in content_type and not url.endswith('.pdf'):
            print(f"⚠️ PDF가 아닌 파일 형식: {content_type}")
            return None
        
        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.pdf', delete=False) as tmp_file:
            tmp_file.write(response.content)
            pdf_path = tmp_file.name
        
        print(f"✅ PDF 다운로드 완료: {pdf_path}")
        
        # 캐시 저장
        if cache_key:
            cache_manager.save_pdf_cache(cache_key, pdf_path)
        
        return pdf_path
    
    except Exception as e:
        print(f"❌ PDF 다운로드 실패: {e}")
        return None


def extract_text_from_pdf(pdf_path: str) -> Optional[str]:
    """
    PDF 파일에서 텍스트를 추출합니다.
    
    Args:
        pdf_path: PDF 파일 경로
    
    Returns:
        추출된 텍스트 (실패 시 None)
    """
    if not os.path.exists(pdf_path):
        return None
    
    try:
        print(f"📄 PDF 텍스트 추출 중: {pdf_path}")
        loader = PyPDFLoader(pdf_path)
        documents = loader.load()
        
        # 모든 페이지의 텍스트 결합
        text = "\n\n".join([doc.page_content for doc in documents])
        
        print(f"✅ PDF 텍스트 추출 완료: {len(text)}자")
        return text
    
    except Exception as e:
        print(f"❌ PDF 텍스트 추출 실패: {e}")
        return None


def summarize_pdf_content(text: str, content_type: str = "주의사항", max_length: int = 2000) -> str:
    """
    PDF에서 추출한 긴 텍스트를 ChatGPT로 요약합니다.
    
    Args:
        text: 원본 텍스트
        content_type: 내용 유형 (주의사항, 효능, 사용법 등)
        max_length: 요약된 텍스트의 최대 길이 (자)
    
    Returns:
        요약된 텍스트
    """
    if not text or len(text) < 500:  # 너무 짧으면 요약 불필요
        return text
    
    try:
        print(f"📝 {content_type} 내용 요약 중... (원본: {len(text)}자)")
        
        summary_prompt = f"""당신은 의약품 정보 전문가입니다. 다음 {content_type} 내용을 요약해주세요.

**요약 원칙:**
1. 모든 중요한 정보를 포함하되, 핵심만 간결하게 정리
2. 금기사항, 주의사항, 부작용 등은 반드시 포함
3. 구체적인 수치나 용량 정보는 유지
4. 중복되는 내용은 제거
5. 요약된 내용은 {max_length}자 이내로 작성

**원본 내용:**
{text}

**요약된 {content_type}:**
"""
        
        # 순환 import 방지를 위해 직접 LLM 호출
        from cache_manager import cache_manager
        
        # 캐시 확인
        cached_response = cache_manager.get_llm_response_cache(summary_prompt, "pdf_summary")
        if cached_response:
            summarized = cached_response
        else:
            llm = _get_summary_llm()
            response = llm.invoke(summary_prompt)
            summarized = response.content if hasattr(response, 'content') else str(response)
            # 캐시 저장
            if summarized and len(summarized) > 100:
                cache_manager.save_llm_response_cache(summary_prompt, summarized, "pdf_summary")
        
        if summarized and len(summarized) > 100:
            print(f"✅ 요약 완료: {len(summarized)}자 (원본: {len(text)}자)")
            return summarized
        else:
            print(f"⚠️ 요약 결과가 너무 짧아 원본 사용")
            # 원본이 너무 길면 앞부분만 반환
            return text[:max_length] + "..." if len(text) > max_length else text
    
    except Exception as e:
        print(f"⚠️ 요약 실패, 원본 사용: {e}")
        # 원본이 너무 길면 앞부분만 반환
        return text[:max_length] + "..." if len(text) > max_length else text


def get_pdf_content_from_excel_link(excel_file_path: str, row_index: int, 
                                    simple_name: str, column_mapping: Dict[str, str] = None,
                                    summarize: bool = True, max_length: int = 2000,
                                    hyperlinks: Dict[str, Optional[str]] = None) -> Optional[str]:
    """
    Excel 파일의 특정 셀에서 PDF 링크를 추출하고 내용을 가져옵니다.
    
    Args:
        excel_file_path: Excel 파일 경로
        row_index: 행 인덱스 (0부터 시작, pandas iterrows 기준)
        simple_name: 간단한 이름 (예: '효능', '복용법', '주의사항')
        column_mapping: {간단한_이름: 실제_컬럼명} 매핑
        summarize: 요약 여부 (기본값: True)
        max_length: 요약 시 최대 길이 (기본값: 2000자)
        hyperlinks: 미리 추출한 하이퍼링크 딕셔너리 (선택사항, 제공 시 중복 추출 방지)
    
    Returns:
        PDF에서 추출한 텍스트 내용 (요약된 경우 요약본, 실패 시 None)
    """
    # 하이퍼링크 추출 (제공되지 않은 경우에만)
    if hyperlinks is None:
        hyperlinks = extract_hyperlinks_from_excel(excel_file_path, row_index, column_mapping)
    url = hyperlinks.get(simple_name)
    
    if not url:
        print(f"  ⚠️ {simple_name} 하이퍼링크에서 URL을 찾을 수 없음")
        return None
    
    print(f"  🔗 {simple_name} URL 발견: {url[:80]}...")
    
    # PDF 다운로드
    cache_key = f"pdf_{os.path.basename(excel_file_path)}_{row_index}_{simple_name}"
    pdf_path = download_pdf_from_url(url, cache_key)
    
    if not pdf_path:
        print(f"  ❌ {simple_name} PDF 다운로드 실패")
        return None
    
    print(f"  📄 {simple_name} PDF 다운로드 완료: {os.path.basename(pdf_path)}")
    
    # 텍스트 추출
    text = extract_text_from_pdf(pdf_path)
    
    if not text:
        return None
    
    # 요약 여부 확인
    if summarize and len(text) > 1000:  # 1000자 이상이면 요약
        text = summarize_pdf_content(text, content_type=simple_name, max_length=max_length)
    
    return text


def enrich_excel_row_with_pdf_content(excel_file_path: str, row_index: int, 
                                     link_columns: List[str] = None,
                                     column_mapping: Dict[str, str] = None,
                                     summarize: bool = True, max_length: int = 2000) -> Dict[str, Optional[str]]:
    """
    Excel 행의 링크 컬럼들에서 PDF 내용을 추출하여 반환합니다. (병렬 처리 버전)
    
    Args:
        excel_file_path: Excel 파일 경로
        row_index: 행 인덱스 (0부터 시작, pandas iterrows 기준)
        link_columns: PDF 링크가 있는 컬럼명 리스트 (기본값: ['효능', '복용법', '주의사항'])
        column_mapping: {간단한_이름: 실제_컬럼명} 매핑
        summarize: 요약 여부 (기본값: True)
        max_length: 요약 시 최대 길이 (기본값: 2000자)
    
    Returns:
        {간단한_이름: PDF 내용} 딕셔너리
    """
    print(f"📥 PDF 다운로드 시작 (병렬 처리): {os.path.basename(excel_file_path)}, 행 {row_index}")
    
    if link_columns is None:
        link_columns = ['효능', '복용법', '주의사항']
    
    # 먼저 모든 하이퍼링크를 한 번에 추출
    hyperlinks = extract_hyperlinks_from_excel(excel_file_path, row_index, column_mapping)
    
    # 병렬로 PDF 처리
    def process_single_pdf(simple_name: str) -> tuple:
        """단일 PDF 처리 함수 (병렬 처리용)"""
        try:
            print(f"🔍 {simple_name} PDF 처리 시작...")
            # 미리 추출한 하이퍼링크 전달하여 중복 추출 방지
            content = get_pdf_content_from_excel_link(
                excel_file_path, row_index, simple_name, column_mapping, summarize, max_length, hyperlinks
            )
            if content:
                print(f"✅ {simple_name} PDF 내용 추출 완료 (길이: {len(content)}자)")
            else:
                print(f"⚠️ {simple_name} PDF 내용 추출 실패 또는 URL 없음")
            return (simple_name, content)
        except Exception as e:
            print(f"❌ {simple_name} PDF 처리 중 오류: {e}")
            return (simple_name, None)
    
    # 병렬 실행
    pdf_contents = {}
    with ThreadPoolExecutor(max_workers=len(link_columns)) as executor:
        # 모든 PDF 처리 작업 제출
        futures = {
            executor.submit(process_single_pdf, simple_name): simple_name 
            for simple_name in link_columns
        }
        
        # 결과 수집
        for future in as_completed(futures):
            try:
                simple_name, content = future.result()
                pdf_contents[simple_name] = content
            except Exception as e:
                simple_name = futures[future]
                print(f"❌ {simple_name} PDF 처리 실패: {e}")
                pdf_contents[simple_name] = None
    
    print(f"📤 PDF 다운로드 완료: {len([k for k, v in pdf_contents.items() if v])}개 성공")
    return pdf_contents

